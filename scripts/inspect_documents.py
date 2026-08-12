"""Helper utilities to inspect a single PDF, CV (PDF), or XLSX file to guide manual eyeballing.

Usage examples:
    python scripts/inspect_documents.py pdf path/to/file.pdf
    python scripts/inspect_documents.py cv path/to/cv.pdf
    python scripts/inspect_documents.py xlsx path/to/file.xlsx
"""
import sys
import re
from pathlib import Path

def inspect_pdf(path):
    import pdfplumber
    p = Path(path)
    if not p.exists():
        print('File not found:', path)
        return
    with pdfplumber.open(path) as pdf:
        text = '\n'.join(page.extract_text() or '' for page in pdf.pages[:5])
    print('--- SAMPLE TEXT (first pages) ---')
    print(text[:2000])
    print('\n--- CURRENCY HINTS FOUND ---')
    for m in re.finditer(r'Rs\.?\s?[\d,]+(?:\.\d+)?|INR\s?[\d,]+|[\d,]+\s?Cr', text):
        print('-', m.group(0))
    print('\n--- POSSIBLE CLIENT NAMES / SIGNATURE BLOCKS ---')
    for line in text.splitlines():
        if 'client' in line.lower() or 'for' in line.lower() and len(line) < 100:
            print('-', line.strip())

def inspect_cv(path):
    # CVs often have the candidate name at top and project lists; extract top lines
    import pdfplumber
    p = Path(path)
    with pdfplumber.open(path) as pdf:
        first_page = pdf.pages[0]
        text = first_page.extract_text() or ''
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    print('--- TOP LINES ---')
    for ln in lines[:12]:
        print('-', ln)
    print('\nLook for project bullets, years, and employer names to associate with the top name.')

def inspect_xlsx(path):
    import openpyxl
    from openpyxl import load_workbook
    p = Path(path)
    if not p.exists():
        print('File not found:', path)
        return
    wb = load_workbook(path, data_only=True)
    print('Sheets:', wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    print('Sample cells (A1:C10):')
    for r in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=3, values_only=True):
        print(r)

def usage():
    print('Usage: python scripts/inspect_documents.py [pdf|cv|xlsx] path')

def main():
    if len(sys.argv) < 3:
        usage(); return
    kind = sys.argv[1]
    path = sys.argv[2]
    if kind == 'pdf':
        inspect_pdf(path)
    elif kind == 'cv':
        inspect_cv(path)
    elif kind == 'xlsx':
        inspect_xlsx(path)
    else:
        usage()

if __name__ == '__main__':
    main()
